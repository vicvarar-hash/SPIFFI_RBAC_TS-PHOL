"""Add the '5.3.4 LLM-authored rules' tab to PALADIN_experiments_v2.xlsx.

Results for the three retained authors (Opus 4.8, GPT-5.4, Gemini 3.1 Pro) plus
PALADIN production, with LIVE formulas so every metric is provably derived from
the confusion matrix, followed by the exact prompt/brief shared with the models.
Confusion matrices are taken verbatim from scratch/authored_results.json.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = r"C:\Users\vivargas\.copilot\repos\copilot-worktrees\PALADIN_RESEARCH\vivargas-microsoft-effective-tribble"
XLSX = ROOT + r"\reports\PALADIN_experiments_v2.xlsx"
BRIEF = ROOT + r"\policies\llm_authored\AUTHORING_BRIEF.md"
RESULTS = ROOT + r"\scratch\authored_results.json"
SHEET = "5.3.4 LLM-authored rules"

# ---- palette (matches existing tabs) ----
DARK = "FF1F3864"      # title / header fill
LIGHT = "FFDDEBF7"     # emphasis / section fill
GREY = "FF595959"      # notes
WHITE = "FFFFFFFF"
HFILL = PatternFill("solid", fgColor=DARK)
LFILL = PatternFill("solid", fgColor=LIGHT)

def style(c, *, val=None, bold=False, size=11, color=None, fill=None,
          wrap=False, align=None, valign=None, numfmt=None, font="Calibri"):
    if val is not None:
        c.value = val
    c.font = Font(name=font, bold=bold, size=size, color=color)
    if fill is not None:
        c.fill = fill
    c.alignment = Alignment(wrap_text=wrap, horizontal=align, vertical=valign)
    if numfmt:
        c.number_format = numfmt
    return c

# ---- load confusion matrices ----
with open(RESULTS, encoding="utf-8") as f:
    res = {r["name"]: r for r in json.load(f)}

# (label, key, is_prod)
ROWS = [
    ("Opus 4.8  (strictest)", "opus48", False),
    ("GPT-5.4", "gpt54", False),
    ("Gemini 3.1 Pro  (most permissive)", "gemini31", False),
    ("PALADIN production", "PROD_FILES", True),
]

wb = openpyxl.load_workbook(XLSX)
if SHEET in wb.sheetnames:
    del wb[SHEET]
# insert right before "Logs on GitHub"
idx = wb.sheetnames.index("Logs on GitHub")
ws = wb.create_sheet(SHEET, index=idx)

widths = {"A": 26, "B": 7, "C": 7, "D": 8, "E": 8, "F": 14,
          "G": 14, "H": 13, "I": 13, "J": 15, "K": 15}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ---- title + notes ----
style(ws["A1"], val="5.3.4  -  Best-effort adversarial rule design  (LLM-authored RBAC + ABAC)",
      bold=True, size=14, color=DARK)
notes = [
    "Table 12 in the paper. Three frontier LLMs (one per vendor) each INDEPENDENTLY authored a complete RBAC+ABAC policy from a leak-free brief; TRAC is frozen at production. Each authored policy is evaluated through the identical PALADIN stack.",
    "Validation mode, n = 6,942 (1,157 tasks x 6 personas). Model-independent. Domain-eligible (valid) = 1,977 = TP+FN;  not-valid = 4,965 = FP+TN  -  both constant across every policy, so all rows are comparable.",
    "Derivations (LIVE formulas in the cells): Retention R_ret = TP/(TP+FN);   Security-failure (leak) = FP/(FP+TN);   TRAC drop-one (pp) = (leak_FULL - leak_noTRAC) x 100  (negative = leak RISES when TRAC is removed).",
    "Provenance: scratch/run_authored_experiment.py replaying datasets/llm_inference_logs/20260708132606_claude-opus-4-8_validation.json; raw confusion matrices in scratch/authored_results.json; authored policies in policies/llm_authored/<model>/{rbac.yaml, abac_rules.yaml}.",
]
for i, t in enumerate(notes, start=2):
    style(ws[f"A{i}"], val=t, size=10, color=GREY, wrap=True)
    ws.merge_cells(f"A{i}:K{i}")

# ---- results header ----
hdr = ["Policy author (RBAC+ABAC)", "TP", "FP", "FN", "TN",
       "Retention R_ret", "Sec-fail (FULL)", "FP (no TRAC)", "TN (no TRAC)",
       "Sec-fail (no TRAC)", "TRAC drop-one"]
HR = 7
for j, h in enumerate(hdr, start=1):
    style(ws.cell(HR, j), val=h, bold=True, color=WHITE, fill=HFILL,
          wrap=True, align="center", valign="center")
ws.row_dimensions[HR].height = 30

# ---- data rows with live formulas ----
for k, (label, key, is_prod) in enumerate(ROWS):
    r = HR + 1 + k
    m = res[key]["FULL"]
    nt = res[key]["noTRAC"]
    fill = LFILL if is_prod else None
    bold = is_prod
    style(ws.cell(r, 1), val=label, bold=bold, fill=fill)
    for col, v in ((2, m["TP"]), (3, m["FP"]), (4, m["FN"]), (5, m["TN"]),
                   (8, nt["FP"]), (9, nt["TN"])):
        style(ws.cell(r, col), val=v, bold=bold, fill=fill, align="center")
    style(ws.cell(r, 6), val=f"=B{r}/(B{r}+D{r})", bold=bold, fill=fill,
          align="center", numfmt="0.0%")
    style(ws.cell(r, 7), val=f"=C{r}/(C{r}+E{r})", bold=bold, fill=fill,
          align="center", numfmt="0.0%")
    style(ws.cell(r, 10), val=f"=H{r}/(H{r}+I{r})", bold=bold, fill=fill,
          align="center", numfmt="0.0%")
    style(ws.cell(r, 11), val=f"=(G{r}-J{r})*100", bold=bold, fill=fill,
          align="center", numfmt='0.0"pp"')

# ---- interpretation ----
IR = HR + 1 + len(ROWS) + 1  # blank line then block
style(ws.cell(IR, 1), val="What it shows", bold=True, size=11, color=DARK, fill=LFILL)
ws.merge_cells(start_row=IR, start_column=1, end_row=IR, end_column=11)
reads = [
    "No authored conventional stack (no-TRAC, column J) reaches PALADIN's 0.103 floor. Best = Opus at 0.190 = 1.85x the floor, bought only by collapsing retention to 0.340 (FULL) / 0.389 (no-TRAC). PALADIN's full stack PARETO-DOMINATES all three authors on BOTH axes at once (higher retention AND lower leak).",
    "TRAC's drop-one is strictly negative on every policy (-9.5 to -18.9 pp) and grows MONOTONICALLY with permissiveness. Panel-mean no-TRAC leak = 26.3%; TRAC removes 14.2 pp (> half). Conclusion: better conventional rules shift the security/retention frontier but cannot cross it -- TRAC stays load-bearing.",
]
for i, t in enumerate(reads):
    rr = IR + 1 + i
    style(ws.cell(rr, 1), val=t, size=10, color=GREY, wrap=True)
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=11)
    ws.row_dimensions[rr].height = 42

# ---- prompt section ----
PR = IR + 1 + len(reads) + 1
style(ws.cell(PR, 1), val="PROMPT  /  AUTHORING BRIEF SHARED WITH EACH MODEL",
      bold=True, size=11, color=DARK, fill=LFILL)
ws.merge_cells(start_row=PR, start_column=1, end_row=PR, end_column=11)

wrapper_intro = ("Each model ran independently as a background policy-architect (one per vendor), given ONLY the brief below: "
                 "no repository access, no task texts, no ground-truth labels, no persona-domain pairings, and none of PALADIN's shipped rules. Wrapper instruction sent to each model:")
style(ws.cell(PR + 1, 1), val=wrapper_intro, size=10, color=GREY, wrap=True)
ws.merge_cells(start_row=PR + 1, start_column=1, end_row=PR + 1, end_column=11)
ws.row_dimensions[PR + 1].height = 42

wrapper = (
    "You are a senior access-control engineer. Read ONLY the authoring brief below and design the best least-privilege "
    "RBAC + ABAC policy you can for the six agent personas and nine MCP tool domains it describes. Do not open or infer any "
    "other file in the repository. You have no task texts, no ground-truth labels, and no list of which persona-domain "
    "pairings are correct.\n"
    "Emit EXACTLY two fenced YAML blocks -- first the complete rbac.yaml, then the complete abac_rules.yaml -- matching the "
    "schemas in the brief, each with a <=5-line design-rationale comment block inside the YAML. Optimise to deny 'wrong' "
    "(right server, wrong tools) and 'null' (wrong server) bundles while admitting legitimate work; do not blanket-deny. "
    "A separate fixed task-relevance layer runs after yours -- do not try to model it."
)
style(ws.cell(PR + 2, 1), val=wrapper, size=10, font="Consolas", wrap=True, valign="top")
ws.merge_cells(start_row=PR + 2, start_column=1, end_row=PR + 2, end_column=11)
ws.row_dimensions[PR + 2].height = 150

# ---- full brief, verbatim, one line per row (monospace) ----
BH = PR + 4
style(ws.cell(BH, 1), val="FULL AUTHORING BRIEF  (verbatim; also at policies/llm_authored/AUTHORING_BRIEF.md)",
      bold=True, size=11, color=DARK, fill=LFILL)
ws.merge_cells(start_row=BH, start_column=1, end_row=BH, end_column=11)

with open(BRIEF, encoding="utf-8") as f:
    brief_lines = f.read().splitlines()
start = BH + 1
for i, line in enumerate(brief_lines):
    style(ws.cell(start + i, 1), val=line, size=9, font="Consolas", color="FF333333")

wb.save(XLSX)
print(f"OK: wrote sheet '{SHEET}' at index {idx}; {len(brief_lines)} brief lines; sheets now:")
print(openpyxl.load_workbook(XLSX).sheetnames)
